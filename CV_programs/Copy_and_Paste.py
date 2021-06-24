import openpyxl
import os
from CV_programs import paste1,graph,graph1
def copy(file_name,Sheet,area,ands):


    Accumulation=1
    p=0

    wb=openpyxl.Workbook()
    wb.save('data.xlsx')
    datawb = openpyxl.load_workbook(filename='data.xlsx')
    dataws = datawb.active
    dataws.cell(row = 3, column = 1, value = Accumulation )
    datawb.save('data.xlsx')


    i = 0
    while i < 3:
        print("source file\'s delimiter[space:1 tab;2 comma;3 excel;4]")
        de=input()

        if de.isdecimal():

            delimiter=int(de)
            if delimiter==1 or delimiter==2 or delimiter==3 or delimiter==4:
                break

    while i < 3:
        print("The number of Column of Capacity\'s value")
        Ca=input()
        if Ca.isdecimal():
            break
    Capacity=int(Ca)

    while i < 3:
        print("The number of Column of Voltage\'s value")
        Vol=input()
        if Vol.isdecimal():
            break
    Voltage=int(Vol)



    while i<2:
        print("frequency [If you want to go next sheet,write \"exit\" and press Enter]")
        frequency=input()
        if(p>3):
            frequency="exit"


        if frequency=="exit":
            if p==0:
                break
            else:
                if delimiter==1 or delimiter==2 or delimiter==3:
                    os.remove('test.xlsx')

                outwb = openpyxl.load_workbook(filename=file_name)
                # Select the Sheet（active is the sheet cyrrently open）
                outws=outwb[Sheet]
                #Fill in the cell with area
                outws['A3']= "S[cm^2]"
                outws['A4']= area
                outwb.save(file_name)
                break
        else:

            txt="Drag and Drop "+str(frequency)+"'s file here"
            print(txt)
            source=input()
            print("Please waite for a white")
            paste1.paste(file_name,Sheet,frequency,source,delimiter,Accumulation,Capacity,p)

            datawb = openpyxl.load_workbook(filename='data.xlsx')
            dataws = datawb.active
            ABCD=dataws.cell(row = 1, column = p+1 ).value
            Accumulation=Accumulation+int(ABCD)+3
            p+=1

            dataws.cell(row = 3, column = p+1, value = Accumulation )
            dataws.cell(row = 4, column = p, value = frequency )
            datawb.save('data.xlsx')

    dataws.cell(row = 5, column = 1 ,value = Voltage)
    V=dataws.cell(row = 5, column = 1 ).value


    graph.add_to_chart(Sheet,file_name,p,Voltage)
    graph1.add_to_chart(Sheet,file_name,p,Voltage,Capacity)

    os.remove('data.xlsx')
